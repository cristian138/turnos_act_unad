from fastapi import FastAPI, APIRouter, Depends, HTTPException, status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
import socketio
import os
import logging
import json
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from jose import JWTError, jwt
from passlib.context import CryptContext
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from fastapi.responses import StreamingResponse
from contextlib import asynccontextmanager
import aiomysql
from uuid import uuid4

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Configuración MySQL
MYSQL_HOST = os.environ.get('MYSQL_HOST', 'localhost')
MYSQL_PORT = int(os.environ.get('MYSQL_PORT', 3306))
MYSQL_USER = os.environ.get('MYSQL_USER', 'root')
MYSQL_PASSWORD = os.environ.get('MYSQL_PASSWORD', '')
MYSQL_DATABASE = os.environ.get('MYSQL_DATABASE', 'turnos_unad')

SECRET_KEY = os.environ.get('SECRET_KEY', 'tu-clave-secreta-super-segura-cambiala-en-produccion')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 480

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

# Pool de conexiones MySQL
pool = None

async def get_pool():
    global pool
    if pool is None:
        pool = await aiomysql.create_pool(
            host=MYSQL_HOST,
            port=MYSQL_PORT,
            user=MYSQL_USER,
            password=MYSQL_PASSWORD,
            db=MYSQL_DATABASE,
            autocommit=True,
            charset='utf8mb4'
        )
    return pool

async def execute_query(query: str, params: tuple = None, fetch: str = None):
    """Ejecuta una query y retorna resultados"""
    p = await get_pool()
    async with p.acquire() as conn:
        async with conn.cursor(aiomysql.DictCursor) as cur:
            await cur.execute(query, params)
            if fetch == 'one':
                return await cur.fetchone()
            elif fetch == 'all':
                return await cur.fetchall()
            return cur.lastrowid

sio = socketio.AsyncServer(async_mode='asgi', cors_allowed_origins='*')
app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ============== Modelos Pydantic ==============

class Usuario(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    nombre: str
    email: EmailStr
    rol: str
    activo: bool = True
    servicios_asignados: List[str] = []
    modulo: Optional[str] = None
    fecha_creacion: str

class UsuarioCreate(BaseModel):
    nombre: str
    email: EmailStr
    password: str
    rol: str
    servicios_asignados: List[str] = []
    modulo: Optional[str] = None

class UsuarioUpdate(BaseModel):
    nombre: Optional[str] = None
    email: Optional[EmailStr] = None
    password: Optional[str] = None
    rol: Optional[str] = None
    activo: Optional[bool] = None
    servicios_asignados: Optional[List[str]] = None
    modulo: Optional[str] = None

class LoginRequest(BaseModel):
    email: EmailStr
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str
    usuario: Usuario

class Servicio(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    nombre: str
    prefijo: str
    activo: bool = True
    fecha_creacion: str

class ServicioCreate(BaseModel):
    nombre: str
    prefijo: str

class ServicioUpdate(BaseModel):
    nombre: Optional[str] = None
    prefijo: Optional[str] = None
    activo: Optional[bool] = None

class Turno(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    codigo: str
    servicio_id: str
    servicio_nombre: str
    prioridad: Optional[str] = None
    estado: str
    observaciones: Optional[str] = None
    funcionario_id: Optional[str] = None
    funcionario_nombre: Optional[str] = None
    modulo: Optional[str] = None
    fecha_creacion: str
    fecha_llamado: Optional[str] = None
    fecha_atencion: Optional[str] = None
    fecha_cierre: Optional[str] = None
    tiempo_espera: Optional[int] = None
    tiempo_atencion: Optional[int] = None
    tipo_documento: str
    numero_documento: str
    nombre_completo: str
    telefono: str
    correo: EmailStr
    tipo_usuario: str

class TurnoCreate(BaseModel):
    servicio_id: str
    prioridad: Optional[str] = None
    observaciones: Optional[str] = None
    tipo_documento: str
    numero_documento: str
    nombre_completo: str
    telefono: str
    correo: EmailStr
    tipo_usuario: str

class TurnoLlamar(BaseModel):
    turno_id: str
    modulo: Optional[str] = None

class TurnoAtender(BaseModel):
    turno_id: str

class TurnoCerrar(BaseModel):
    turno_id: str

class TurnoRedirigir(BaseModel):
    turno_id: str
    nuevo_servicio_id: str

class Configuracion(BaseModel):
    model_config = ConfigDict(extra="ignore")
    impresion_habilitada: bool = True
    prioridades: List[str] = ["Discapacidad", "Embarazo", "Adulto Mayor"]

class ConfiguracionUpdate(BaseModel):
    impresion_habilitada: Optional[bool] = None
    prioridades: Optional[List[str]] = None

# ============== Funciones de Autenticación ==============

def verificar_password(password_plano: str, password_hash: str) -> bool:
    return pwd_context.verify(password_plano, password_hash)

def obtener_password_hash(password: str) -> str:
    return pwd_context.hash(password)

def crear_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def row_to_usuario(row: dict) -> Usuario:
    """Convierte una fila de BD a modelo Usuario"""
    servicios = json.loads(row.get('servicios_asignados') or '[]')
    return Usuario(
        id=row['id'],
        nombre=row['nombre'],
        email=row['email'],
        rol=row['rol'],
        activo=bool(row['activo']),
        servicios_asignados=servicios,
        modulo=row.get('modulo'),
        fecha_creacion=row['fecha_creacion'].isoformat() if isinstance(row['fecha_creacion'], datetime) else row['fecha_creacion']
    )

def row_to_servicio(row: dict) -> Servicio:
    """Convierte una fila de BD a modelo Servicio"""
    return Servicio(
        id=row['id'],
        nombre=row['nombre'],
        prefijo=row['prefijo'],
        activo=bool(row['activo']),
        fecha_creacion=row['fecha_creacion'].isoformat() if isinstance(row['fecha_creacion'], datetime) else row['fecha_creacion']
    )

def row_to_turno(row: dict) -> Turno:
    """Convierte una fila de BD a modelo Turno"""
    def format_date(val):
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.isoformat()
        return val
    
    return Turno(
        id=row['id'],
        codigo=row['codigo'],
        servicio_id=row['servicio_id'],
        servicio_nombre=row['servicio_nombre'],
        prioridad=row.get('prioridad'),
        estado=row['estado'],
        observaciones=row.get('observaciones'),
        funcionario_id=row.get('funcionario_id'),
        funcionario_nombre=row.get('funcionario_nombre'),
        modulo=row.get('modulo'),
        fecha_creacion=format_date(row['fecha_creacion']),
        fecha_llamado=format_date(row.get('fecha_llamado')),
        fecha_atencion=format_date(row.get('fecha_atencion')),
        fecha_cierre=format_date(row.get('fecha_cierre')),
        tiempo_espera=row.get('tiempo_espera'),
        tiempo_atencion=row.get('tiempo_atencion'),
        tipo_documento=row['tipo_documento'],
        numero_documento=row['numero_documento'],
        nombre_completo=row['nombre_completo'],
        telefono=row['telefono'],
        correo=row['correo'],
        tipo_usuario=row['tipo_usuario']
    )

async def obtener_usuario_actual(credentials: HTTPAuthorizationCredentials = Depends(security)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="No se pudieron validar las credenciales",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        email: str = payload.get("sub")
        if email is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    
    row = await execute_query(
        "SELECT * FROM usuarios WHERE email = %s",
        (email,),
        fetch='one'
    )
    if row is None:
        raise credentials_exception
    return row_to_usuario(row)

def requerir_rol(roles_permitidos: List[str]):
    async def verificar_rol(usuario: Usuario = Depends(obtener_usuario_actual)):
        if usuario.rol not in roles_permitidos:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=f"Rol '{usuario.rol}' no tiene permiso para esta acción"
            )
        return usuario
    return verificar_rol

# ============== Endpoints de Autenticación ==============

@api_router.post("/auth/login", response_model=Token)
async def login(datos: LoginRequest):
    row = await execute_query(
        "SELECT * FROM usuarios WHERE email = %s",
        (datos.email,),
        fetch='one'
    )
    
    if not row or not verificar_password(datos.password, row['password_hash']):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Email o contraseña incorrectos"
        )
    
    if not row['activo']:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Usuario desactivado"
        )
    
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = crear_access_token(
        data={"sub": row['email']}, expires_delta=access_token_expires
    )
    
    usuario = row_to_usuario(row)
    return Token(access_token=access_token, token_type="bearer", usuario=usuario)

@api_router.get("/auth/me", response_model=Usuario)
async def obtener_perfil(usuario: Usuario = Depends(obtener_usuario_actual)):
    return usuario

# ============== Endpoints de Usuarios ==============

@api_router.get("/usuarios", response_model=List[Usuario])
async def listar_usuarios(usuario: Usuario = Depends(requerir_rol(["administrador"]))):
    rows = await execute_query("SELECT * FROM usuarios ORDER BY fecha_creacion DESC", fetch='all')
    return [row_to_usuario(row) for row in rows]

@api_router.post("/usuarios", response_model=Usuario)
async def crear_usuario(datos: UsuarioCreate, usuario: Usuario = Depends(requerir_rol(["administrador"]))):
    existing = await execute_query(
        "SELECT id FROM usuarios WHERE email = %s",
        (datos.email,),
        fetch='one'
    )
    if existing:
        raise HTTPException(status_code=400, detail="El email ya está registrado")
    
    usuario_id = str(uuid4())
    password_hash = obtener_password_hash(datos.password)
    fecha_creacion = datetime.now(timezone.utc)
    servicios_json = json.dumps(datos.servicios_asignados)
    
    await execute_query(
        """INSERT INTO usuarios (id, nombre, email, password_hash, rol, activo, servicios_asignados, modulo, fecha_creacion)
           VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
        (usuario_id, datos.nombre, datos.email, password_hash, datos.rol, True, servicios_json, datos.modulo, fecha_creacion)
    )
    
    row = await execute_query("SELECT * FROM usuarios WHERE id = %s", (usuario_id,), fetch='one')
    return row_to_usuario(row)

@api_router.put("/usuarios/{usuario_id}", response_model=Usuario)
async def actualizar_usuario(usuario_id: str, datos: UsuarioUpdate, usuario: Usuario = Depends(requerir_rol(["administrador"]))):
    existing = await execute_query("SELECT * FROM usuarios WHERE id = %s", (usuario_id,), fetch='one')
    if not existing:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    updates = []
    params = []
    
    if datos.nombre is not None:
        updates.append("nombre = %s")
        params.append(datos.nombre)
    if datos.email is not None:
        updates.append("email = %s")
        params.append(datos.email)
    if datos.password is not None:
        updates.append("password_hash = %s")
        params.append(obtener_password_hash(datos.password))
    if datos.rol is not None:
        updates.append("rol = %s")
        params.append(datos.rol)
    if datos.activo is not None:
        updates.append("activo = %s")
        params.append(datos.activo)
    if datos.servicios_asignados is not None:
        updates.append("servicios_asignados = %s")
        params.append(json.dumps(datos.servicios_asignados))
    if datos.modulo is not None:
        updates.append("modulo = %s")
        params.append(datos.modulo)
    
    if updates:
        params.append(usuario_id)
        await execute_query(
            f"UPDATE usuarios SET {', '.join(updates)} WHERE id = %s",
            tuple(params)
        )
    
    row = await execute_query("SELECT * FROM usuarios WHERE id = %s", (usuario_id,), fetch='one')
    return row_to_usuario(row)

@api_router.delete("/usuarios/{usuario_id}")
async def eliminar_usuario(usuario_id: str, usuario: Usuario = Depends(requerir_rol(["administrador"]))):
    existing = await execute_query("SELECT * FROM usuarios WHERE id = %s", (usuario_id,), fetch='one')
    if not existing:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    await execute_query("DELETE FROM usuarios WHERE id = %s", (usuario_id,))
    return {"mensaje": "Usuario eliminado exitosamente"}

# ============== Endpoints de Servicios ==============

@api_router.get("/servicios", response_model=List[Servicio])
async def listar_servicios(usuario: Usuario = Depends(obtener_usuario_actual)):
    rows = await execute_query("SELECT * FROM servicios ORDER BY nombre", fetch='all')
    return [row_to_servicio(row) for row in rows]

@api_router.post("/servicios", response_model=Servicio)
async def crear_servicio(datos: ServicioCreate, usuario: Usuario = Depends(requerir_rol(["administrador"]))):
    servicio_id = str(uuid4())
    fecha_creacion = datetime.now(timezone.utc)
    
    await execute_query(
        """INSERT INTO servicios (id, nombre, prefijo, activo, fecha_creacion)
           VALUES (%s, %s, %s, %s, %s)""",
        (servicio_id, datos.nombre, datos.prefijo.upper(), True, fecha_creacion)
    )
    
    row = await execute_query("SELECT * FROM servicios WHERE id = %s", (servicio_id,), fetch='one')
    return row_to_servicio(row)

@api_router.put("/servicios/{servicio_id}", response_model=Servicio)
async def actualizar_servicio(servicio_id: str, datos: ServicioUpdate, usuario: Usuario = Depends(requerir_rol(["administrador"]))):
    existing = await execute_query("SELECT * FROM servicios WHERE id = %s", (servicio_id,), fetch='one')
    if not existing:
        raise HTTPException(status_code=404, detail="Servicio no encontrado")
    
    updates = []
    params = []
    
    if datos.nombre is not None:
        updates.append("nombre = %s")
        params.append(datos.nombre)
    if datos.prefijo is not None:
        updates.append("prefijo = %s")
        params.append(datos.prefijo.upper())
    if datos.activo is not None:
        updates.append("activo = %s")
        params.append(datos.activo)
    
    if updates:
        params.append(servicio_id)
        await execute_query(
            f"UPDATE servicios SET {', '.join(updates)} WHERE id = %s",
            tuple(params)
        )
    
    row = await execute_query("SELECT * FROM servicios WHERE id = %s", (servicio_id,), fetch='one')
    return row_to_servicio(row)

@api_router.delete("/servicios/{servicio_id}")
async def eliminar_servicio(servicio_id: str, usuario: Usuario = Depends(requerir_rol(["administrador"]))):
    existing = await execute_query("SELECT * FROM servicios WHERE id = %s", (servicio_id,), fetch='one')
    if not existing:
        raise HTTPException(status_code=404, detail="Servicio no encontrado")
    
    await execute_query("DELETE FROM servicios WHERE id = %s", (servicio_id,))
    return {"mensaje": "Servicio eliminado exitosamente"}

# ============== Endpoints de Turnos ==============

@api_router.post("/turnos/generar", response_model=Turno)
async def generar_turno(datos: TurnoCreate, usuario: Usuario = Depends(obtener_usuario_actual)):
    servicio = await execute_query("SELECT * FROM servicios WHERE id = %s", (datos.servicio_id,), fetch='one')
    if not servicio:
        raise HTTPException(status_code=404, detail="Servicio no encontrado")
    
    # Obtener el último número del día para este servicio
    hoy = datetime.now(timezone.utc).date()
    row = await execute_query(
        """SELECT codigo FROM turnos 
           WHERE servicio_id = %s AND DATE(fecha_creacion) = %s 
           ORDER BY codigo DESC LIMIT 1""",
        (datos.servicio_id, hoy),
        fetch='one'
    )
    
    if row:
        ultimo_num = int(row['codigo'].split('-')[1])
        nuevo_num = ultimo_num + 1
    else:
        nuevo_num = 1
    
    codigo = f"{servicio['prefijo']}-{nuevo_num:03d}"
    turno_id = str(uuid4())
    fecha_creacion = datetime.now(timezone.utc)
    
    # Guardar o actualizar cliente
    cliente_existente = await execute_query(
        "SELECT id FROM clientes WHERE numero_documento = %s",
        (datos.numero_documento,),
        fetch='one'
    )
    
    if cliente_existente:
        await execute_query(
            """UPDATE clientes SET nombre_completo = %s, telefono = %s, correo = %s, 
               tipo_usuario = %s, tipo_documento = %s WHERE numero_documento = %s""",
            (datos.nombre_completo, datos.telefono, datos.correo, datos.tipo_usuario, datos.tipo_documento, datos.numero_documento)
        )
    else:
        await execute_query(
            """INSERT INTO clientes (id, tipo_documento, numero_documento, nombre_completo, telefono, correo, tipo_usuario)
               VALUES (%s, %s, %s, %s, %s, %s, %s)""",
            (str(uuid4()), datos.tipo_documento, datos.numero_documento, datos.nombre_completo, datos.telefono, datos.correo, datos.tipo_usuario)
        )
    
    # Crear turno
    await execute_query(
        """INSERT INTO turnos (id, codigo, servicio_id, servicio_nombre, prioridad, estado, observaciones,
           fecha_creacion, tipo_documento, numero_documento, nombre_completo, telefono, correo, tipo_usuario)
           VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
        (turno_id, codigo, datos.servicio_id, servicio['nombre'], datos.prioridad, 'creado', datos.observaciones,
         fecha_creacion, datos.tipo_documento, datos.numero_documento, datos.nombre_completo, datos.telefono, datos.correo, datos.tipo_usuario)
    )
    
    row = await execute_query("SELECT * FROM turnos WHERE id = %s", (turno_id,), fetch='one')
    turno = row_to_turno(row)
    
    await sio.emit('nuevo_turno', turno.model_dump())
    
    return turno

@api_router.get("/turnos/cola/{servicio_id}", response_model=List[Turno])
async def obtener_cola_servicio(servicio_id: str, usuario: Usuario = Depends(obtener_usuario_actual)):
    rows = await execute_query(
        """SELECT * FROM turnos WHERE servicio_id = %s AND estado = 'creado' 
           ORDER BY prioridad IS NULL, fecha_creacion ASC""",
        (servicio_id,),
        fetch='all'
    )
    return [row_to_turno(row) for row in rows]

@api_router.get("/turnos/todos", response_model=List[Turno])
async def obtener_todos_turnos(usuario: Usuario = Depends(obtener_usuario_actual)):
    if usuario.rol == "funcionario":
        servicios_ids = usuario.servicios_asignados
        if not servicios_ids:
            return []
        placeholders = ','.join(['%s'] * len(servicios_ids))
        rows = await execute_query(
            f"""SELECT * FROM turnos WHERE servicio_id IN ({placeholders}) AND estado = 'creado' 
               ORDER BY prioridad IS NULL, fecha_creacion ASC""",
            tuple(servicios_ids),
            fetch='all'
        )
    else:
        rows = await execute_query(
            """SELECT * FROM turnos WHERE estado = 'creado' 
               ORDER BY prioridad IS NULL, fecha_creacion ASC""",
            fetch='all'
        )
    return [row_to_turno(row) for row in rows]

@api_router.get("/turnos/lista-completa", response_model=List[Turno])
async def obtener_lista_completa_turnos(usuario: Usuario = Depends(obtener_usuario_actual)):
    if usuario.rol not in ["vap", "administrador"]:
        raise HTTPException(status_code=403, detail="No tienes permisos para ver esta lista")
    
    hoy = datetime.now(timezone.utc).date()
    rows = await execute_query(
        """SELECT * FROM turnos WHERE DATE(fecha_creacion) = %s 
           ORDER BY fecha_creacion DESC""",
        (hoy,),
        fetch='all'
    )
    return [row_to_turno(row) for row in rows]

@api_router.post("/turnos/cancelar", response_model=Turno)
async def cancelar_turno_pendiente(datos: TurnoCerrar, usuario: Usuario = Depends(requerir_rol(["administrador"]))):
    turno = await execute_query("SELECT * FROM turnos WHERE id = %s", (datos.turno_id,), fetch='one')
    if not turno:
        raise HTTPException(status_code=404, detail="Turno no encontrado")
    
    if turno['estado'] != 'creado':
        raise HTTPException(status_code=400, detail="Solo se pueden cancelar turnos en estado pendiente (creado)")
    
    fecha_cierre = datetime.now(timezone.utc)
    await execute_query(
        """UPDATE turnos SET estado = 'cancelado', fecha_cierre = %s, 
           funcionario_id = %s, funcionario_nombre = %s WHERE id = %s""",
        (fecha_cierre, usuario.id, f"Cancelado por: {usuario.nombre}", datos.turno_id)
    )
    
    row = await execute_query("SELECT * FROM turnos WHERE id = %s", (datos.turno_id,), fetch='one')
    turno_actualizado = row_to_turno(row)
    
    await sio.emit('turno_cancelado', turno_actualizado.model_dump())
    return turno_actualizado

@api_router.post("/turnos/llamar", response_model=Turno)
async def llamar_turno(datos: TurnoLlamar, usuario: Usuario = Depends(requerir_rol(["funcionario", "administrador"]))):
    turno = await execute_query("SELECT * FROM turnos WHERE id = %s", (datos.turno_id,), fetch='one')
    if not turno:
        raise HTTPException(status_code=404, detail="Turno no encontrado")
    
    if turno['estado'] != 'creado':
        raise HTTPException(status_code=400, detail="El turno no está en estado creado")
    
    if usuario.rol == "funcionario" and turno['servicio_id'] not in usuario.servicios_asignados:
        raise HTTPException(status_code=403, detail="No tienes asignado este servicio")
    
    fecha_llamado = datetime.now(timezone.utc)
    fecha_creacion = turno['fecha_creacion']
    if isinstance(fecha_creacion, str):
        fecha_creacion = datetime.fromisoformat(fecha_creacion.replace('Z', '+00:00'))
    tiempo_espera = int((fecha_llamado - fecha_creacion).total_seconds())
    
    modulo_asignado = usuario.modulo or datos.modulo or f"Módulo {usuario.nombre.split()[0]}"
    
    await execute_query(
        """UPDATE turnos SET estado = 'llamado', funcionario_id = %s, funcionario_nombre = %s,
           modulo = %s, fecha_llamado = %s, tiempo_espera = %s WHERE id = %s""",
        (usuario.id, usuario.nombre, modulo_asignado, fecha_llamado, tiempo_espera, datos.turno_id)
    )
    
    row = await execute_query("SELECT * FROM turnos WHERE id = %s", (datos.turno_id,), fetch='one')
    turno_actualizado = row_to_turno(row)
    
    await sio.emit('turno_llamado', turno_actualizado.model_dump())
    return turno_actualizado

@api_router.post("/turnos/atender", response_model=Turno)
async def atender_turno(datos: TurnoAtender, usuario: Usuario = Depends(requerir_rol(["funcionario", "administrador"]))):
    turno = await execute_query("SELECT * FROM turnos WHERE id = %s", (datos.turno_id,), fetch='one')
    if not turno:
        raise HTTPException(status_code=404, detail="Turno no encontrado")
    
    if turno['estado'] != 'llamado':
        raise HTTPException(status_code=400, detail="El turno no está en estado llamado")
    
    fecha_atencion = datetime.now(timezone.utc)
    
    await execute_query(
        "UPDATE turnos SET estado = 'atendiendo', fecha_atencion = %s WHERE id = %s",
        (fecha_atencion, datos.turno_id)
    )
    
    row = await execute_query("SELECT * FROM turnos WHERE id = %s", (datos.turno_id,), fetch='one')
    turno_actualizado = row_to_turno(row)
    
    await sio.emit('turno_atendiendo', turno_actualizado.model_dump())
    return turno_actualizado

@api_router.post("/turnos/cerrar", response_model=Turno)
async def cerrar_turno(datos: TurnoCerrar, usuario: Usuario = Depends(requerir_rol(["funcionario", "administrador"]))):
    turno = await execute_query("SELECT * FROM turnos WHERE id = %s", (datos.turno_id,), fetch='one')
    if not turno:
        raise HTTPException(status_code=404, detail="Turno no encontrado")
    
    if turno['estado'] not in ['llamado', 'atendiendo']:
        raise HTTPException(status_code=400, detail="El turno no puede ser cerrado en su estado actual")
    
    fecha_cierre = datetime.now(timezone.utc)
    fecha_atencion = turno.get('fecha_atencion') or turno.get('fecha_llamado')
    if isinstance(fecha_atencion, str):
        fecha_atencion = datetime.fromisoformat(fecha_atencion.replace('Z', '+00:00'))
    tiempo_atencion = int((fecha_cierre - fecha_atencion).total_seconds()) if fecha_atencion else 0
    
    await execute_query(
        "UPDATE turnos SET estado = 'finalizado', fecha_cierre = %s, tiempo_atencion = %s WHERE id = %s",
        (fecha_cierre, tiempo_atencion, datos.turno_id)
    )
    
    row = await execute_query("SELECT * FROM turnos WHERE id = %s", (datos.turno_id,), fetch='one')
    turno_actualizado = row_to_turno(row)
    
    await sio.emit('turno_finalizado', turno_actualizado.model_dump())
    return turno_actualizado

@api_router.post("/turnos/redirigir", response_model=Turno)
async def redirigir_turno(datos: TurnoRedirigir, usuario: Usuario = Depends(requerir_rol(["funcionario", "administrador"]))):
    turno = await execute_query("SELECT * FROM turnos WHERE id = %s", (datos.turno_id,), fetch='one')
    if not turno:
        raise HTTPException(status_code=404, detail="Turno no encontrado")
    
    nuevo_servicio = await execute_query("SELECT * FROM servicios WHERE id = %s", (datos.nuevo_servicio_id,), fetch='one')
    if not nuevo_servicio:
        raise HTTPException(status_code=404, detail="Servicio destino no encontrado")
    
    # Generar nuevo código
    hoy = datetime.now(timezone.utc).date()
    row = await execute_query(
        """SELECT codigo FROM turnos 
           WHERE servicio_id = %s AND DATE(fecha_creacion) = %s 
           ORDER BY codigo DESC LIMIT 1""",
        (datos.nuevo_servicio_id, hoy),
        fetch='one'
    )
    
    if row:
        ultimo_num = int(row['codigo'].split('-')[1])
        nuevo_num = ultimo_num + 1
    else:
        nuevo_num = 1
    
    nuevo_codigo = f"{nuevo_servicio['prefijo']}-{nuevo_num:03d}"
    
    await execute_query(
        """UPDATE turnos SET servicio_id = %s, servicio_nombre = %s, codigo = %s, estado = 'creado',
           funcionario_id = NULL, funcionario_nombre = NULL, modulo = NULL, fecha_llamado = NULL WHERE id = %s""",
        (datos.nuevo_servicio_id, nuevo_servicio['nombre'], nuevo_codigo, datos.turno_id)
    )
    
    row = await execute_query("SELECT * FROM turnos WHERE id = %s", (datos.turno_id,), fetch='one')
    turno_actualizado = row_to_turno(row)
    
    await sio.emit('turno_redirigido', turno_actualizado.model_dump())
    return turno_actualizado

@api_router.get("/turnos/llamados-recientes")
async def obtener_llamados_recientes():
    rows = await execute_query(
        """SELECT * FROM turnos WHERE estado IN ('llamado', 'atendiendo') 
           ORDER BY fecha_llamado DESC LIMIT 10""",
        fetch='all'
    )
    return [row_to_turno(row).model_dump() for row in rows]

# ============== Endpoint de Clientes ==============

@api_router.get("/clientes/buscar/{numero_documento}")
async def buscar_cliente(numero_documento: str, usuario: Usuario = Depends(obtener_usuario_actual)):
    cliente = await execute_query(
        "SELECT * FROM clientes WHERE numero_documento = %s",
        (numero_documento,),
        fetch='one'
    )
    
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    
    return {
        "tipo_documento": cliente['tipo_documento'],
        "numero_documento": cliente['numero_documento'],
        "nombre_completo": cliente['nombre_completo'],
        "telefono": cliente['telefono'],
        "correo": cliente['correo'],
        "tipo_usuario": cliente['tipo_usuario']
    }

# ============== Endpoints de Configuración ==============

@api_router.get("/configuracion", response_model=Configuracion)
async def obtener_configuracion(usuario: Usuario = Depends(obtener_usuario_actual)):
    row = await execute_query("SELECT * FROM configuracion LIMIT 1", fetch='one')
    if not row:
        return Configuracion()
    
    prioridades = json.loads(row.get('prioridades') or '["Discapacidad", "Embarazo", "Adulto Mayor"]')
    return Configuracion(
        impresion_habilitada=bool(row['impresion_habilitada']),
        prioridades=prioridades
    )

@api_router.put("/configuracion", response_model=Configuracion)
async def actualizar_configuracion(datos: ConfiguracionUpdate, usuario: Usuario = Depends(requerir_rol(["administrador"]))):
    existing = await execute_query("SELECT * FROM configuracion LIMIT 1", fetch='one')
    
    if not existing:
        prioridades_json = json.dumps(datos.prioridades or ["Discapacidad", "Embarazo", "Adulto Mayor"])
        await execute_query(
            "INSERT INTO configuracion (impresion_habilitada, prioridades) VALUES (%s, %s)",
            (datos.impresion_habilitada if datos.impresion_habilitada is not None else True, prioridades_json)
        )
    else:
        updates = []
        params = []
        if datos.impresion_habilitada is not None:
            updates.append("impresion_habilitada = %s")
            params.append(datos.impresion_habilitada)
        if datos.prioridades is not None:
            updates.append("prioridades = %s")
            params.append(json.dumps(datos.prioridades))
        
        if updates:
            await execute_query(f"UPDATE configuracion SET {', '.join(updates)}", tuple(params))
    
    return await obtener_configuracion(usuario)

# ============== Endpoints de Reportes ==============

@api_router.get("/reportes/atencion")
async def generar_reporte_atencion(
    fecha_inicio: str = None,
    fecha_fin: str = None,
    servicio_id: str = None,
    funcionario_id: str = None,
    formato: str = "json",
    usuario: Usuario = Depends(requerir_rol(["administrador"]))
):
    query = "SELECT * FROM turnos WHERE estado = 'finalizado'"
    params = []
    
    if fecha_inicio:
        query += " AND DATE(fecha_creacion) >= %s"
        params.append(fecha_inicio)
    if fecha_fin:
        query += " AND DATE(fecha_creacion) <= %s"
        params.append(fecha_fin)
    if servicio_id:
        query += " AND servicio_id = %s"
        params.append(servicio_id)
    if funcionario_id:
        query += " AND funcionario_id = %s"
        params.append(funcionario_id)
    
    query += " ORDER BY fecha_creacion DESC"
    
    rows = await execute_query(query, tuple(params) if params else None, fetch='all')
    turnos = [row_to_turno(row) for row in rows]
    
    if formato == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Atención"
        
        headers = ["Código", "Servicio", "Cliente", "Documento", "Funcionario", "Módulo", 
                   "Fecha Creación", "Tiempo Espera (min)", "Tiempo Atención (min)", "Estado"]
        
        header_fill = PatternFill(start_color="0A4F77", end_color="0A4F77", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        for row_num, turno in enumerate(turnos, 2):
            ws.cell(row=row_num, column=1, value=turno.codigo)
            ws.cell(row=row_num, column=2, value=turno.servicio_nombre)
            ws.cell(row=row_num, column=3, value=turno.nombre_completo)
            ws.cell(row=row_num, column=4, value=turno.numero_documento)
            ws.cell(row=row_num, column=5, value=turno.funcionario_nombre or "")
            ws.cell(row=row_num, column=6, value=turno.modulo or "")
            ws.cell(row=row_num, column=7, value=turno.fecha_creacion)
            ws.cell(row=row_num, column=8, value=round(turno.tiempo_espera / 60, 2) if turno.tiempo_espera else 0)
            ws.cell(row=row_num, column=9, value=round(turno.tiempo_atencion / 60, 2) if turno.tiempo_atencion else 0)
            ws.cell(row=row_num, column=10, value=turno.estado)
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=reporte_atencion_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )
    
    return {"turnos": [t.model_dump() for t in turnos], "total": len(turnos)}

# ============== WebSocket Events ==============

@sio.event
async def connect(sid, environ):
    logger.info(f"Cliente conectado: {sid}")

@sio.event
async def disconnect(sid):
    logger.info(f"Cliente desconectado: {sid}")

# ============== Health Check ==============

@api_router.get("/health")
async def health_check():
    try:
        await execute_query("SELECT 1", fetch='one')
        return {"status": "ok", "database": "connected"}
    except Exception as e:
        return {"status": "error", "database": str(e)}

# ============== Configuración de la App ==============

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.include_router(api_router)
socket_app = socketio.ASGIApp(sio, other_asgi_app=app)
app = socket_app
